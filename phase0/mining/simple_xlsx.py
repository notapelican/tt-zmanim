"""Build the SIMPLE review workbook from simple_rows.json: friendly wording,
two tabs, yellow answer columns, attention highlighting. No formulas."""
from __future__ import annotations
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

MINE = Path(__file__).parent
data = json.loads((MINE / "simple_rows.json").read_text())

ARIAL = "Arial"
YELLOW = PatternFill("solid", fgColor="FFF2AB")
GREY = PatternFill("solid", fgColor="F2F2F2")
HDR = PatternFill("solid", fgColor="1F3864")
BAND = PatternFill("solid", fgColor="EDF2FA")
ATTN = PatternFill("solid", fgColor="FDE8E8")
THIN = Border(*(Side(style="thin", color="BFBFBF"),) * 4)


def put(ws, row, col, val, *, bold=False, fill=None, color="000000", size=11,
        italic=False, wrap=True, valign="top"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name=ARIAL, bold=bold, color=color, size=size, italic=italic)
    if fill:
        c.fill = fill
    c.alignment = Alignment(wrap_text=wrap, vertical=valign)
    c.border = THIN
    return c


wb = Workbook()
ws = wb.active
ws.title = "Timesheet rules"

put(ws, 1, 1, "Timesheet special-day rules — please review", bold=True, size=14, wrap=False)
put(ws, 2, 1, "Each row describes something the timesheet will print automatically on a "
    "special day. Please fill in the two YELLOW columns: pick an answer from the dropdown, "
    "and add a note if anything needs changing. Rows shaded PINK have a question we "
    "particularly need your help with.", italic=True, wrap=False)
put(ws, 3, 1, "Example answer:  Looks right   |   or:  Needs a change — \"Mincha on a fast "
    "should be 30 minutes before shkia, not 25\"", italic=True, wrap=False, color="666666")

HEADERS = ["#", "Section", "When it appears", "What the sheet will show",
           "Example from a past sheet", "Question for you", "Your answer", "Your notes"]
HR = 5
for j, h in enumerate(HEADERS, 1):
    put(ws, HR, j, h, bold=True, fill=HDR, color="FFFFFF", valign="center")

row = HR + 1
last_sec = None
band = False
for r in data["rules"]:
    if r["section"] != last_sec:
        band = not band
        last_sec = r["section"]
    fill = ATTN if r.get("attention") else (BAND if band else None)
    vals = [r["id"], r["section"], r["when"], r["shows"], r.get("example", ""),
            r.get("question", "")]
    for j, v in enumerate(vals, 1):
        put(ws, row, j, v, fill=fill)
    put(ws, row, 7, "", fill=YELLOW)
    put(ws, row, 8, "", fill=YELLOW)
    row += 1

dv = DataValidation(type="list",
                    formula1='"Looks right,Needs a change (see notes),Leave it out,Not sure"',
                    allow_blank=True, showDropDown=False)
ws.add_data_validation(dv)
dv.add(f"G{HR + 1}:G{row - 1}")

for j, w in enumerate([7, 18, 28, 52, 38, 36, 20, 36], 1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = f"A{HR + 1}"
ws.auto_filter.ref = f"A{HR}:H{row - 1}"

# --- questions tab ---
ws2 = wb.create_sheet("Questions")
put(ws2, 1, 1, "Questions we need answered", bold=True, size=14, wrap=False)
put(ws2, 2, 1, "These came up while studying the old sheets. Please answer in the YELLOW "
    "column, in your own words.", italic=True, wrap=False)
H2 = ["#", "Topic", "Question", "Your answer"]
for j, h in enumerate(H2, 1):
    put(ws2, 4, j, h, bold=True, fill=HDR, color="FFFFFF", valign="center")
r2 = 5
for i, q in enumerate(data["questions"], 1):
    put(ws2, r2, 1, i)
    put(ws2, r2, 2, q["topic"])
    put(ws2, r2, 3, q["question"])
    put(ws2, r2, 4, "", fill=YELLOW)
    r2 += 1
for j, w in enumerate([5, 26, 75, 55], 1):
    ws2.column_dimensions[get_column_letter(j)].width = w
ws2.freeze_panes = "A5"

out = MINE / "TTCC-special-days-review-SIMPLE.xlsx"
wb.save(out)
print("wrote", out, "| rules:", len(data["rules"]), "| questions:", len(data["questions"]))
