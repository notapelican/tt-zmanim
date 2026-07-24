"""Convert SPECIAL-DAYS-CATALOG.md into a review spreadsheet with yellow
Decision/Notes columns the shul fills in. No formulas — pure data + formatting."""
from __future__ import annotations
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

MINE = Path(__file__).parent
MD = (MINE / "SPECIAL-DAYS-CATALOG.md").read_text()

ARIAL = "Arial"
YELLOW = PatternFill("solid", fgColor="FFF2AB")
GREY = PatternFill("solid", fgColor="F2F2F2")
HDR_FILL = PatternFill("solid", fgColor="1F3864")
BAND = PatternFill("solid", fgColor="EDF2FA")
THIN = Border(*(Side(style="thin", color="BFBFBF"),) * 4)
CONF_COLOR = {"HIGH": "1A7F37", "MEDIUM": "B45309", "LOW": "B91C1C"}


def clean(s: str) -> str:
    s = re.sub(r"\*\*(.+?)\*\*", r"\1", s)
    s = s.replace("**", "").replace("`", "")
    return re.sub(r"[ \t]+", " ", s).strip()


def flatten_block(lines: list[str]) -> str:
    """Markdown block (may contain a table) -> readable multi-line cell text."""
    out = []
    for ln in lines:
        t = ln.strip()
        if not t:
            continue
        if re.match(r"^\|[\s:|-]+\|$", t):  # table separator row
            continue
        if t.startswith("|"):
            cells = [clean(c) for c in t.strip("|").split("|")]
            out.append("  ".join(c for c in cells if c) if any(out) else " | ".join(cells))
            continue
        out.append(clean(t.lstrip("-•").strip()))
    return "\n".join(out)


# --- parse rules -----------------------------------------------------------
cat_re = re.compile(r"^## (\d+)\. (.+)$")
rule_re = re.compile(r"^### Rule (\S+) — (.+)$")
key_re = re.compile(r"^- \*\*(Trigger|Emits|Evidence|Confidence|Open question)([^*]*)\*\*[^:]*:\s*(.*)$")
other_bullet_re = re.compile(r"^- \*\*(.+?)\*\*:?\s*(.*)$")

rules = []
cat = None
cur = None
key = None
for ln in MD.splitlines():
    m = cat_re.match(ln)
    if m:
        cat = f"{m.group(1)}. {clean(m.group(2))}"
        continue
    if ln.startswith("## "):  # appendix sections end rule parsing
        cat = None
        if cur:
            rules.append(cur); cur = None
        continue
    m = rule_re.match(ln)
    if m:
        if cur:
            rules.append(cur)
        cur = {"id": m.group(1), "name": clean(m.group(2)), "cat": cat,
               "Trigger": [], "Emits": [], "Evidence": [], "Confidence": [],
               "Open question": []}
        key = None
        continue
    if cur is None:
        continue
    m = key_re.match(ln)
    if m:
        key = m.group(1)
        variant = m.group(2).strip()
        content = (f"({variant.strip('()')}) " if variant else "") + m.group(3)
        if content.strip():
            cur[key].append(content)
        continue
    m = other_bullet_re.match(ln)
    if m and cur is not None and not ln.startswith("  "):
        # A bold bullet with a nonstandard key (e.g. "**Not present**"):
        # fold it into Emits so nothing is silently dropped.
        key = "Emits"
        cur[key].append(f"{m.group(1)}: {m.group(2)}" if m.group(2) else m.group(1))
        continue
    if ln.strip() == "---":
        key = None
        continue
    if key:
        cur[key].append(ln)
if cur:
    rules.append(cur)

# --- parse appendices ------------------------------------------------------
def section(md, start, end=None):
    pat = re.escape(start)
    m = re.search(rf"^## {pat}.*?$", md, re.M)
    if not m:
        return ""
    rest = md[m.end():]
    m2 = re.search(r"^## ", rest, re.M)
    return rest[: m2.start()] if m2 else rest


def numbered_items(text):
    items, cur = [], None
    for ln in text.splitlines():
        m = re.match(r"^(\d+)\.\s+(.*)$", ln.strip())
        if m:
            if cur:
                items.append(cur)
            cur = m.group(2)
        elif cur is not None and ln.strip():
            cur += " " + ln.strip()
        elif cur is not None and not ln.strip():
            items.append(cur); cur = None
    if cur:
        items.append(cur)
    return [clean(i) for i in items]


def bullet_items(text):
    items, cur = [], None
    for ln in text.splitlines():
        s = ln.strip()
        if s.startswith("- "):
            if cur:
                items.append(cur)
            cur = s[2:]
        elif cur is not None and s:
            cur += " " + s
        elif cur is not None:
            items.append(cur); cur = None
    if cur:
        items.append(cur)
    return [clean(i) for i in items]


errors = numbered_items(section(MD, "Suspected sheet errors found"))
unexplained = numbered_items(section(MD, "Unexplained items needing shul input"))
gaps = bullet_items(section(MD, "Coverage gaps"))

print(f"parsed: {len(rules)} rules, {len(errors)} errors, "
      f"{len(unexplained)} unexplained, {len(gaps)} gaps")

# --- build workbook --------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Rules"

def put(ws, row, col, val, *, bold=False, fill=None, color="000000", size=10,
        italic=False, wrap=True, valign="top"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name=ARIAL, bold=bold, color=color, size=size, italic=italic)
    if fill:
        c.fill = fill
    c.alignment = Alignment(wrap_text=wrap, vertical=valign)
    c.border = THIN
    return c

# Legend
put(ws, 1, 1, "TTCC Zmanim — Special-day automation rules: REVIEW SHEET",
    bold=True, size=13, wrap=False)
put(ws, 2, 1, "Fill in the two YELLOW columns only. Decision: pick from the dropdown "
    "(Confirm / Correct (see notes) / Remove / Unsure). Your notes: corrections in plain "
    "English. The EXAMPLE row below shows the expected format — ignore/overwrite it.",
    italic=True, wrap=False)

HEADERS = ["Rule ID", "Category", "Rule", "Trigger (when it fires)",
           "Generates (lines/notes + time formula)", "Evidence from your sheets",
           "Confidence", "Open question", "Decision", "Your notes"]
HR = 4
for j, h in enumerate(HEADERS, 1):
    put(ws, HR, j, h, bold=True, fill=HDR_FILL, color="FFFFFF", valign="center")

# Example row
ex = ["1c", "1. Minor fasts", "EXAMPLE — Fast-day afternoon Mincha",
      "minor fast day (weekday)", "Fast-day Mincha line, shkia − 25 min",
      "(see real row below)", "HIGH", "",
      "Correct (see notes)", "Mincha should be shkia − 30, not − 25"]
for j, v in enumerate(ex, 1):
    put(ws, HR + 1, j, v, italic=True, fill=GREY)

row = HR + 2
last_cat = None
band = False
for r in rules:
    if r["cat"] != last_cat:
        band = not band
        last_cat = r["cat"]
    fill = BAND if band else None
    conf_text = flatten_block(r["Confidence"])
    conf_key = next((k for k in ("HIGH", "MEDIUM", "LOW") if k in conf_text.upper()), "")
    trigger = flatten_block(r["Trigger"]) or "(implicit in the rule name)"
    vals = [r["id"], r["cat"], r["name"], trigger,
            flatten_block(r["Emits"]), flatten_block(r["Evidence"]),
            conf_text, flatten_block(r["Open question"])]
    for j, v in enumerate(vals, 1):
        c = put(ws, row, j, v, fill=fill)
        if j == 7 and conf_key:
            c.font = Font(name=ARIAL, bold=True, color=CONF_COLOR[conf_key], size=10)
    put(ws, row, 9, "", fill=YELLOW)
    put(ws, row, 10, "", fill=YELLOW)
    row += 1

dv = DataValidation(type="list",
                    formula1='"Confirm,Correct (see notes),Remove,Unsure"',
                    allow_blank=True, showDropDown=False)
ws.add_data_validation(dv)
dv.add(f"I{HR + 1}:I{row - 1}")

widths = [8, 22, 30, 32, 48, 48, 22, 36, 18, 40]
for j, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = f"A{HR + 1}"
ws.auto_filter.ref = f"A{HR}:J{row - 1}"

# --- second sheet: errors / questions / gaps --------------------------------
ws2 = wb.create_sheet("Errors & questions")
put(ws2, 1, 1, "Suspected sheet errors, unexplained items, and coverage gaps",
    bold=True, size=13, wrap=False)
put(ws2, 2, 1, "Fill in the YELLOW column: confirm/deny errors, answer questions, and "
    "describe the intended behaviour for coverage gaps.", italic=True, wrap=False)
H2 = ["#", "Type", "Item", "Your answer / notes"]
for j, h in enumerate(H2, 1):
    put(ws2, 4, j, h, bold=True, fill=HDR_FILL, color="FFFFFF", valign="center")
r2 = 5
for typ, arr in [("Suspected error", errors),
                 ("Unexplained item — needs your input", unexplained),
                 ("Coverage gap — supply intended behaviour", gaps)]:
    for i, item in enumerate(arr, 1):
        put(ws2, r2, 1, i)
        put(ws2, r2, 2, typ)
        put(ws2, r2, 3, item)
        put(ws2, r2, 4, "", fill=YELLOW)
        r2 += 1
for j, w in enumerate([5, 32, 90, 45], 1):
    ws2.column_dimensions[get_column_letter(j)].width = w
ws2.freeze_panes = "A5"

out = MINE / "TTCC-special-days-review.xlsx"
wb.save(out)
print("wrote", out)
